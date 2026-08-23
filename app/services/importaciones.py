"""Importación segura y confirmada de catálogos e inventario."""

from __future__ import annotations

import csv
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
import re
import unicodedata
from zipfile import BadZipFile, ZipFile

from flask import current_app
from itsdangerous import BadSignature, SignatureExpired, URLSafeTimedSerializer
from openpyxl import load_workbook
from pypdf import PdfReader

from ..models import Producto, db
from ..permisos import evaluar_permiso
from .auditoria import registrar_auditoria
from .contexto import ContextoOperacion
from .inventario import ServicioInventario
from .productos import ServicioProductos


class ErrorImportacion(ValueError):
    codigo = "importacion_invalida"


ALIASES = {
    "codigo": "codigo",
    "sku": "codigo",
    "codigo_producto": "codigo",
    "nombre": "nombre",
    "producto": "nombre",
    "descripcion": "descripcion",
    "codigo_barras": "codigo_barras",
    "barcode": "codigo_barras",
    "ean": "codigo_barras",
    "categoria": "categoria",
    "subcategoria": "subcategoria",
    "marca": "marca",
    "unidad": "unidad_medida",
    "unidad_medida": "unidad_medida",
    "costo": "costo_referencia",
    "costo_unitario": "costo_referencia",
    "precio": "precio_venta",
    "precio_venta": "precio_venta",
    "stock": "stock_inicial",
    "stock_inicial": "stock_inicial",
    "stock_minimo": "stock_minimo",
    "punto_reorden": "punto_reorden",
    "stock_maximo": "stock_maximo",
}

NUMERICOS = {
    "costo_referencia",
    "precio_venta",
    "stock_inicial",
    "stock_minimo",
    "punto_reorden",
    "stock_maximo",
}


def _normalizar(texto):
    limpio = unicodedata.normalize("NFKD", str(texto or ""))
    limpio = "".join(c for c in limpio if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", "_", limpio.lower()).strip("_")


def _valor(valor):
    if valor is None:
        return ""
    return str(valor).strip()


class ServicioImportaciones:
    def __init__(self, usuario, contexto: ContextoOperacion):
        self.usuario = usuario
        self.contexto = contexto
        if usuario.empresa_id != contexto.empresa_id:
            raise PermissionError("El contexto no pertenece al usuario")

    def previsualizar(self, archivo):
        self._exigir()
        if not archivo or not archivo.filename:
            raise ErrorImportacion("Selecciona un archivo")
        nombre = archivo.filename.rsplit("/", 1)[-1].rsplit("\\", 1)[-1]
        extension = nombre.rsplit(".", 1)[-1].lower() if "." in nombre else ""
        if extension not in {"csv", "xlsx", "pdf"}:
            raise ErrorImportacion("Formato no admitido; usa CSV, XLSX o PDF")
        contenido = archivo.read(current_app.config["IMPORTACION_MAX_BYTES"] + 1)
        if not contenido:
            raise ErrorImportacion("El archivo está vacío")
        if len(contenido) > current_app.config["IMPORTACION_MAX_BYTES"]:
            raise ErrorImportacion("El archivo supera el tamaño máximo permitido")
        if extension == "csv":
            filas = self._csv(contenido)
        elif extension == "xlsx":
            filas = self._xlsx(contenido)
        else:
            filas = self._pdf(contenido)
        resultado = self._preparar(filas)
        token = self._firmador().dumps(
            {
                "empresa_id": self.usuario.empresa_id,
                "usuario_id": self.usuario.id,
                "filas": resultado,
            }
        )
        return {
            "archivo": nombre,
            "formato": extension,
            "filas": resultado,
            "total": len(resultado),
            "validas": sum(not f["errores"] for f in resultado),
            "con_errores": sum(bool(f["errores"]) for f in resultado),
            "token": token,
        }

    def confirmar(self, token):
        self._exigir()
        if not isinstance(token, str) or not token:
            raise ErrorImportacion("Falta la confirmación de la vista previa")
        try:
            carga = self._firmador().loads(token, max_age=1800)
        except SignatureExpired as exc:
            raise ErrorImportacion("La vista previa venció; carga el archivo nuevamente") from exc
        except BadSignature as exc:
            raise ErrorImportacion("La confirmación de importación no es válida") from exc
        if (
            carga.get("empresa_id") != self.usuario.empresa_id
            or carga.get("usuario_id") != self.usuario.id
        ):
            raise PermissionError("La importación no pertenece al usuario")
        filas = carga.get("filas") or []
        if not filas or any(f.get("errores") for f in filas):
            raise ErrorImportacion("Corrige todas las filas antes de importar")
        productos = ServicioProductos(self.usuario)
        creados = actualizados = movimientos = 0
        try:
            for fila in filas:
                datos = dict(fila["datos"])
                stock = Decimal(datos.pop("stock_inicial", "0") or "0")
                existente = db.session.scalar(
                    db.select(Producto).where(
                        Producto.empresa_id == self.usuario.empresa_id,
                        Producto.codigo == datos["codigo"],
                        Producto.eliminado.is_(False),
                    )
                )
                if existente:
                    producto = productos.editar(existente.id, confirmar=False, **datos)
                    actualizados += 1
                else:
                    producto = productos.crear(confirmar=False, **datos)
                    creados += 1
                if stock > 0:
                    ServicioInventario(self.usuario, self.contexto).entrada(
                        producto_id=producto.id,
                        cantidad=stock,
                        costo_unitario=producto.costo_referencia,
                        motivo="Stock inicial importado",
                        referencia_tipo="importacion",
                        confirmar=False,
                    )
                    movimientos += 1
            registrar_auditoria(
                accion="productos.importados",
                modulo="productos",
                usuario_id=self.usuario.id,
                empresa_id=self.usuario.empresa_id,
                entidad_tipo="Importacion",
                datos_nuevos={
                    "creados": creados,
                    "actualizados": actualizados,
                    "movimientos": movimientos,
                },
            )
            db.session.commit()
        except Exception:
            db.session.rollback()
            raise
        return {"creados": creados, "actualizados": actualizados, "movimientos_stock": movimientos}

    def _preparar(self, filas):
        maximo = current_app.config["IMPORTACION_MAX_FILAS"]
        if not filas:
            raise ErrorImportacion("No se encontraron filas importables")
        if len(filas) > maximo:
            raise ErrorImportacion(f"El archivo supera el máximo de {maximo} filas")
        resultado = []
        vistos = set()
        for numero, origen in enumerate(filas, start=2):
            datos = {
                ALIASES.get(_normalizar(k)): _valor(v)
                for k, v in origen.items()
                if ALIASES.get(_normalizar(k))
            }
            datos = {k: v for k, v in datos.items() if v != ""}
            errores = []
            codigo = datos.get("codigo", "").upper()
            nombre = datos.get("nombre", "")
            datos["codigo"] = codigo
            if not codigo:
                errores.append("Falta código")
            if not nombre:
                errores.append("Falta nombre")
            if codigo in vistos:
                errores.append("Código repetido en el archivo")
            vistos.add(codigo)
            for campo in NUMERICOS:
                if campo not in datos:
                    continue
                try:
                    numero_decimal = Decimal(datos[campo].replace(" ", "").replace(",", "."))
                    if numero_decimal < 0:
                        raise InvalidOperation
                    datos[campo] = str(numero_decimal)
                except (InvalidOperation, AttributeError):
                    errores.append(f"{campo.replace('_', ' ')} no es válido")
            resultado.append({"fila": numero, "datos": datos, "errores": errores})
        return resultado

    @staticmethod
    def _csv(contenido):
        try:
            texto = contenido.decode("utf-8-sig")
        except UnicodeDecodeError:
            texto = contenido.decode("latin-1")
        muestra = texto[:4096]
        try:
            dialecto = csv.Sniffer().sniff(muestra, delimiters=",;\t|")
        except csv.Error:
            dialecto = csv.excel
        return list(csv.DictReader(StringIO(texto), dialect=dialecto))

    @staticmethod
    def _xlsx(contenido):
        try:
            with ZipFile(BytesIO(contenido)) as archivo:
                if sum(i.file_size for i in archivo.infolist()) > 50 * 1024 * 1024:
                    raise ErrorImportacion("El Excel descomprimido es demasiado grande")
            libro = load_workbook(BytesIO(contenido), read_only=True, data_only=True)
        except (BadZipFile, KeyError, ValueError) as exc:
            raise ErrorImportacion("El archivo Excel está dañado o no es válido") from exc
        hoja = libro.active
        iterador = hoja.iter_rows(values_only=True)
        encabezados = next(iterador, None)
        if not encabezados:
            return []
        return [
            dict(zip(encabezados, fila))
            for fila in iterador
            if any(v not in (None, "") for v in fila)
        ]

    @staticmethod
    def _pdf(contenido):
        try:
            lector = PdfReader(BytesIO(contenido), strict=False)
            if len(lector.pages) > 30:
                raise ErrorImportacion("El PDF supera el máximo de 30 páginas")
            lineas = []
            for pagina in lector.pages:
                lineas.extend((pagina.extract_text() or "").splitlines())
        except ErrorImportacion:
            raise
        except Exception as exc:
            raise ErrorImportacion(
                "El PDF está protegido, dañado o no contiene texto legible"
            ) from exc
        lineas = [linea.strip() for linea in lineas if linea.strip()]
        if not lineas:
            raise ErrorImportacion(
                "El PDF no contiene texto; los documentos escaneados requieren OCR"
            )
        separador = next((s for s in ("|", ";", "\t") if s in lineas[0]), None)
        usa_espacios = not separador and bool(re.search(r"\s{2,}", lineas[0]))
        if not separador and not usa_espacios:
            raise ErrorImportacion(
                "No se detectó una tabla en el PDF; usa columnas separadas por | o exporta a Excel"
            )
        dividir = (
            (lambda linea: re.split(r"\s{2,}", linea.strip()))
            if usa_espacios
            else (lambda linea: linea.split(separador))
        )
        encabezados = [c.strip() for c in dividir(lineas[0])]
        return [
            dict(zip(encabezados, [c.strip() for c in dividir(linea)]))
            for linea in lineas[1:]
            if len(dividir(linea)) > 1
        ]

    def _firmador(self):
        return URLSafeTimedSerializer(current_app.secret_key, salt="importacion-productos-v1")

    def _exigir(self):
        decision = evaluar_permiso(
            self.usuario, "productos.importar", empresa_id=self.usuario.empresa_id
        )
        if not decision.permitido:
            raise PermissionError(decision.mensaje)
